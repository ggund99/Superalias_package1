# -*- coding: utf-8 -*-
"""
Created on Tue Apr 28 12:38:21 2020

@author: ggund
"""


# -*- coding: utf-8 -*-
"""
Created on Sat Apr 25 10:37:11 2020

@author: ggund
"""
import pandas as pd
import os
from fuzzywuzzy import fuzz
from tqdm import tqdm
import selenium
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import gc
from datetime import datetime
import time
from urllib.parse import urlparse
from tqdm import tqdm
import google
from contextlib import suppress
from googlesearch import search
from openpyxl import workbook
import openpyxl
import re
gc.collect()

class superalias_merge:
      def __init__(self, sleep_time, username  , password, exec_path):
          self.sleep_time = sleep_time
          self.username = username
          self.password = password
          self.exec_path = exec_path
          driver= webdriver.Firefox(executable_path = self.exec_path )
          self.driver = driver
          
          
          
          
      def start_page(self):
          """
          This function opens the Superalias tool in the firefox
          """
          
          self.driver.get('http://tools-p-ec2c.wantedanalytics.com:3001/')
          # Give Username to login the SuperAlias Tool 
          self.driver.find_element_by_xpath('/html/body/form/div[2]/input').send_keys(self.username)
          time.sleep(self.sleep_time)
          # Give password to login to SuperAlias Tool
          self.driver.find_element_by_xpath('/html/body/form/div[3]/input').send_keys(self.password)
          time.sleep(self.sleep_time)
          # clcik on remember password
          if self.driver.find_element_by_xpath('/html/body/form/div[4]/input[2]').is_selected() == True:
             pass
          else:
          # click search in advertisers
             self.driver.find_element_by_xpath('/html/body/form/div[4]/input[2]').click()
             time.sleep(self.sleep_time)
          # Hit the login button to login
          self.driver.find_element_by_xpath('/html/body/form/div[5]/input').click()
          time.sleep(self.sleep_time)
          # Ensure the page is login 
          try:
             element_present = EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr/td[2]/form/div[2]/table/thead/tr/th[5]'))
             WebDriverWait(self.driver, self.sleep_time+5).until(element_present)
             print('Page is Loaded')
          except TimeoutException:
             print('Time out for page loading')
          return('Page is loaded!')
      
      def extract_results(self, query):
          """
          This function extract the results corresponding to the query and create a worksheet

          """
          time.sleep(2)
          t1 = datetime.now()  
          # give location to the chrome driver
          # give url of the search query
          url1 = 'http://tools-p-ec2c.wantedanalytics.com:3001/advertisers?utf8=%E2%9C%93&query='
          url2= '&action_mode=merge&display=superaliases&search_in%5B%5D=advertiser&IsAnonymous=&IsNoise=&IsStaffing2=&limit=&commit=Search'
          self.driver.get(url1+query+url2) 
          with suppress(Exception):
            alert = self.driver.switch_to.alert
            alert.accept()
          time.sleep(2)
          # get number of results
          txt = self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/h2').text
          txt = txt.replace(',', '')
          count = re.findall('\d+', txt)[1]  
          emp_name = query   
          t2 = datetime.now()
          print(query + 'Done')
          return(emp_name, count, t2-t1)
      
        
      def write_results(self, emp_name, count, time, filename):
          wb= openpyxl.Workbook()
          wb.save(filename)
          wb.sheetnames
          sheet = wb.active
          sheet['A1'] = 'Emp_name'
          sheet['B1'] = 'Number_results'
          sheet['C1'] = 'Time'
          row = sheet.max_row+1
          sheet.cell(row = row, column= 1).value = emp_name
          sheet.cell(row = row, column= 2).value = count
          sheet.cell(row = row, column= 3).value = time
          wb.save(filename)
          return(print('records written for ', emp_name))
          
      
            
          
          
      
      def url_extract(self, word):
          """ 
          This function will extract the Top URL from google 
          """
          word = word.replace(' ', '+')
          
          self.driver.get("https://www.google.com/search?safe=active&source=hp&ei=FMmjXua0Go6wrQHmmKvIBg&q="+word+"&oq="+word+"&gs_lcp=CgZwc3ktYWIQAzICCAAyAggAMgIIADICCAAyAggAMgUIABCDATICCAAyBQgAEIMBMgIIADICCAA6DggAEOoCELQCEJoBEOUCUOU7WMpFYKZJaABwAHgAgAH5AogB7geSAQcwLjUuMC4xmAEAoAEBqgEHZ3dzLXdperABBg&sclient=psy-ab&ved=0ahUKEwjmjvDX6oLpAhUOWCsKHWbMCmkQ4dUDCAY&uact=5")
          # get the top URL
          p = self.driver.find_elements_by_css_selector('cite')
          a = ''
          for i,k in enumerate(p):
             if p[i].text != '':
                a = p[i].text
                break
             pass
          # extract the URL from the text
          a = a.split(' ')[0]
          return(a)
      
      
     
          
      def merge(self, query ):
          """
          This function will merge the employers based on similarity of hostnames
          """
          
          # give url of the search query
          url1 = 'http://tools-p-ec2c.wantedanalytics.com:3001/advertisers?utf8=%E2%9C%93&query='
          url2= '&action_mode=merge&display=superaliases&search_in%5B%5D=advertiser&IsAnonymous=&IsNoise=&IsStaffing2=&limit=&commit=Search'
          self.driver.get(url1+query+url2) 
          with suppress(Exception):
             alert = self.driver.switch_to.alert
             alert.accept()

          table = self.driver.find_element_by_xpath("/html/body/table/tbody/tr/td[2]/form/div[2]/table/tbody")
          time.sleep(4)
          if len(table.find_elements_by_tag_name('tr')) > 1:
          # open tab
            self.driver.execute_script("window.open('');")
            # give control to the tab
            self.driver.switch_to.window(self.driver.window_handles[1])
            # Exctract base url
            base_hostname = self.url_extract(word = query)
            # close the tab
            self.driver.close()
            # give control to the superalias window
            self.driver.switch_to.window(self.driver.window_handles[0])
            # get the table of contents from superalias tool
            table = self.driver.find_element_by_xpath("/html/body/table/tbody/tr/td[2]/form/div[2]/table/tbody")
            # get total number of results
            number_of_results = len(table.find_elements_by_tag_name('tr'))
            #loop
            merge_id = ''
            for i in tqdm(range(1,number_of_results+1)):
                time.sleep(2)
            # get each result
                keyword = self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/div[2]/table/tbody/tr[' + str(i)+ ']/td[4]/a').text
                print('keyword obtained  :', keyword)
                # if keyword == query
                if keyword == query:
                    merge_id = self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/div[2]/table/tbody/tr['+ str(i)+ ']/td[2]').text
                else :
                    pass
                # get top url from google
                # open tab
                self.driver.execute_script("window.open('');")
                # give control to the tab
                self.driver.switch_to.window(self.driver.window_handles[1])
                # Exctract base url
                keyword_hostname= self.url_extract(word = keyword)
                # close the tab
                self.driver.close()
                # give the control to superalias tab
                self.driver.switch_to.window(self.driver.window_handles[0])
                print('hostname obtained for  : ', keyword)
                # compare base_hostname with keywordhostname
                time.sleep(2)
                score = fuzz.token_set_ratio(base_hostname, keyword_hostname)
                # select the row only when score > 80
                if score > 80:
                   self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/div[2]/table/tbody/tr['+ str(i) +']/td[1]/input').click()
                   print('Radio button selected for keyword :', keyword)
                else:                            
                   pass
                if i%20== 0:
                   time.sleep( 4)
                   print('sleep for 30 sec : ', i)
                else:
                   pass
                print('selection done for : ' , query)
            # if merge id is obtained then put that else put the first entry as merge_id
                # clear contents of merge field first
            self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/ul/li/input').clear()
            # put merge id in the merge field
            self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/ul/li/input').send_keys(merge_id)
            print('input value in merge box for  ', query)
            # click on the merge button
            self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/ul/li/button').click() 
            print('merge button clicked for ', query)
          else:
            pass
          return(print('Merge done for ', query))
          

